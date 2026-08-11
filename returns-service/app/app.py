"""
QAYDAO Returns Service — isolated sidecar.
- Stores return requests in its OWN postgres DB (`returns`), linked to Chatwoot conversation_id (loose ref).
- Never touches chatwoot_production or Chatwoot tables.
- Serves:
    GET  /returns/api/health
    GET  /returns/api/requests?conversation_id=..   -> latest request for a conversation (for the CS tab)
    GET  /returns/api/requests/{id}
    GET  /returns/api/requests                        -> list (accountant page)
    POST /returns/api/requests                        -> create/update from CS tab
    PATCH /returns/api/requests/{id}/status           -> accountant status change
    GET  /accountant-returns                          -> accountant read-only HTML page (nginx adds basic-auth)
Auth model: the CS tab is only reachable inside Chatwoot (post-login). The accountant page is protected
by nginx basic-auth in front of this service. This service trusts the reverse proxy.
"""
import os
import re
import json
import uuid
import secrets
import asyncpg
import bcrypt
from pathlib import Path
from datetime import date, datetime, timezone, timedelta
from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form, Cookie
from fastapi.responses import JSONResponse, HTMLResponse, Response, FileResponse, RedirectResponse
from pydantic import BaseModel
from typing import Optional

UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "/data/uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
ALLOWED_MIME = {
    "application/pdf": ".pdf",
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
}
SESSION_COOKIE = "returns_session"
SESSION_DAYS_DEFAULT = 1      # normal login
SESSION_DAYS_REMEMBER = 30    # "تذكرني"

DATABASE_URL = os.environ["DATABASE_URL"]
app = FastAPI(title="QAYDAO Returns Service")
_pool: Optional[asyncpg.Pool] = None

REASONS = [
    "المنتج تالف", "المنتج غير مطابق للوصف", "وصل منتج مختلف",
    "العميل غيّر رأيه", "تأخر التوصيل", "مشكلة في المقاس أو اللون",
    "نقص في الطلب", "سبب آخر",
]
ASSIGNEES = ["في", "مروة", "أميرة"]
STATUS_LABELS = {"new": "جديد", "will": "سيتم الإرجاع", "doing": "جاري الإرجاع", "done": "تم الإرجاع", "rejected": "مرفوض", "done_salla": "تم الإرجاع في سلة"}

# --- طريقة الاسترداد (refund method) ---
# tabby/tamara/madfooat => رقم/مرجع واحد. bank_account => البنك + الحساب + الآيبان.
REFUND_LABELS = {
    "tabby": "تابي", "tamara": "تمارا",
    "madfooat": "مدفوع", "bank_account": "الحساب البنكي",
}
REFUND_REF_METHODS = ("tabby", "tamara", "madfooat")
IBAN_RE = re.compile(r"^SA\d{22}$")


def validate_refund(body) -> None:
    """تحقق خادمي (لا يُعتمد على الواجهة وحدها). يرفع 422 عند المخالفة."""
    m = (body.refund_method or "").strip()
    if not m:
        raise HTTPException(422, "طريقة الاسترداد إلزامية.")
    if m not in REFUND_LABELS:
        raise HTTPException(422, "طريقة استرداد غير معروفة.")
    if m in REFUND_REF_METHODS:
        if not (body.refund_reference or "").strip():
            raise HTTPException(422, f"رقم / مرجع {REFUND_LABELS[m]} إلزامي.")
        return
    # bank_account
    if not (body.bank_name or "").strip():
        raise HTTPException(422, "اسم البنك إلزامي.")
    if not (body.bank_account or "").strip():
        raise HTTPException(422, "الحساب البنكي إلزامي.")
    iban = (body.iban or "").strip().upper().replace(" ", "")
    if not IBAN_RE.match(iban):
        raise HTTPException(422, "الآيبان غير صحيح — يجب أن يبدأ بـ SA ويليه 22 رقماً.")
# --- Per-user status permissions (SERVER-enforced, not just UI-hidden) ---
# النذير (procurement) may ONLY set done_salla. Everyone else (financial accountant,
# management) may ONLY set the financial statuses. No overlap between the two roles.
ACCOUNTANT_STATUSES = ["will", "doing", "done", "rejected"]
PURCHASER_STATUSES  = ["done_salla"]
PURCHASER_EMAILS    = {"pr@qaydao.com"}

def allowed_statuses_for(email: str) -> list:
    e = (email or "").lower()
    if e in PURCHASER_EMAILS:
        return list(PURCHASER_STATUSES)
    return list(ACCOUNTANT_STATUSES)


async def pool() -> asyncpg.Pool:
    global _pool
    if _pool is None:
        _pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
    return _pool


@app.on_event("startup")
async def _startup():
    # wait for DB (initdb may still be running)
    import asyncio
    for _ in range(30):
        try:
            p = await pool()
            async with p.acquire() as c:
                await c.execute("SELECT 1")
                await _ensure_auth_tables(c)
            return
        except Exception:
            await asyncio.sleep(1)


async def _ensure_auth_tables(c):
    await c.execute("""
        CREATE TABLE IF NOT EXISTS accountant_users (
            email TEXT PRIMARY KEY, password_hash TEXT NOT NULL,
            display_name TEXT, created_at TIMESTAMPTZ NOT NULL DEFAULT now());
        CREATE TABLE IF NOT EXISTS accountant_sessions (
            token TEXT PRIMARY KEY,
            email TEXT NOT NULL REFERENCES accountant_users(email) ON DELETE CASCADE,
            expires_at TIMESTAMPTZ NOT NULL, created_at TIMESTAMPTZ NOT NULL DEFAULT now());
        CREATE INDEX IF NOT EXISTS idx_sess_expires ON accountant_sessions(expires_at);
    """)
    # opportunistic cleanup of expired sessions
    await c.execute("DELETE FROM accountant_sessions WHERE expires_at < now()")


# ------------------------------- Auth helpers -------------------------------

def _hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def _check_pw(pw: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), h.encode())
    except Exception:
        return False

async def current_user(token: Optional[str]) -> Optional[dict]:
    if not token:
        return None
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow(
            """SELECT s.email, u.display_name FROM accountant_sessions s
                 JOIN accountant_users u ON u.email = s.email
                WHERE s.token=$1 AND s.expires_at > now()""",
            token,
        )
    return dict(r) if r else None

def _require_login_redirect():
    # for HTML pages: send to login
    return RedirectResponse(url="/accountant-login", status_code=302)


def row_to_dict(r) -> dict:
    d = dict(r)
    for k in ("created_at", "updated_at", "contacted_at"):
        if d.get(k) is not None:
            d[k] = d[k].isoformat()
    for k in ("return_created_at", "original_order_at"):
        if d.get(k) is not None:
            d[k] = d[k].isoformat()
    if isinstance(d.get("status_history"), str):
        try:
            d["status_history"] = json.loads(d["status_history"])
        except Exception:
            d["status_history"] = []
    d["status_label"] = STATUS_LABELS.get(d.get("status"), d.get("status"))
    # السجلات القديمة (قبل ميزة طريقة الاسترداد) بنكية ضمنياً — عرضاً فقط، لا يُكتب للـDB.
    m = d.get("refund_method") or "bank_account"
    d["refund_method_label"] = REFUND_LABELS.get(m, m)
    return d


class ReturnIn(BaseModel):
    conversation_id: Optional[int] = None
    customer_name: Optional[str] = None
    order_number: Optional[str] = None
    order_amount: Optional[str] = None
    return_created_at: Optional[str] = None
    original_order_at: Optional[str] = None
    reason: Optional[str] = None
    refund_method: Optional[str] = None
    refund_reference: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    iban: Optional[str] = None
    attachment_name: Optional[str] = None
    assignee: Optional[str] = None
    created_by: Optional[str] = None


class StatusIn(BaseModel):
    status: str
    changed_by: Optional[str] = None
    accountant_note: Optional[str] = None
    reject_reason: Optional[str] = None


# ------------------------------- API -------------------------------

@app.get("/returns/api/health")
async def health():
    try:
        p = await pool()
        async with p.acquire() as c:
            await c.execute("SELECT 1")
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, f"db: {e}")


def _parse_date(v):
    if not v:
        return None
    try:
        return datetime.strptime(v[:10], "%Y-%m-%d").date()
    except Exception:
        return None


@app.post("/returns/api/requests")
async def create_or_update(body: ReturnIn):
    validate_refund(body)
    # تطبيع: كل طريقة تحتفظ بحقولها فقط — لا تُخزَّن بيانات بنكية لطلب تابي/تمارا/مدفوع.
    method = (body.refund_method or "").strip()
    if method in REFUND_REF_METHODS:
        ref = (body.refund_reference or "").strip()
        bank_name = bank_account = iban = None
    else:
        ref = None
        bank_name = (body.bank_name or "").strip()
        bank_account = (body.bank_account or "").strip()
        iban = (body.iban or "").strip().upper().replace(" ", "")
    p = await pool()
    async with p.acquire() as c:
        existing = None
        if body.conversation_id is not None:
            # Only reuse an OPEN request. Closed ones (rejected/done) are final and archived —
            # a re-submission for the same conversation creates a brand-new request.
            existing = await c.fetchrow(
                """SELECT id FROM return_requests
                    WHERE conversation_id=$1 AND status NOT IN ('rejected','done')
                    ORDER BY id DESC LIMIT 1""",
                body.conversation_id,
            )
        rc = _parse_date(body.return_created_at)
        oo = _parse_date(body.original_order_at)
        if existing:
            r = await c.fetchrow(
                """UPDATE return_requests SET
                     customer_name=$2, order_number=$3, order_amount=$4,
                     return_created_at=$5, original_order_at=$6, reason=$7,
                     bank_name=$8, bank_account=$9, iban=$10, attachment_name=$11,
                     assignee=$12, created_by=COALESCE($13, created_by),
                     refund_method=$14, refund_reference=$15
                   WHERE id=$1 RETURNING *""",
                existing["id"], body.customer_name, body.order_number, body.order_amount,
                rc, oo, body.reason, bank_name, bank_account, iban,
                body.attachment_name, body.assignee, body.created_by,
                method, ref,
            )
        else:
            r = await c.fetchrow(
                """INSERT INTO return_requests
                     (conversation_id, customer_name, order_number, order_amount,
                      return_created_at, original_order_at, reason, bank_name,
                      bank_account, iban, attachment_name, assignee, status, created_by,
                      refund_method, refund_reference)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,'new',$13,$14,$15)
                   RETURNING *""",
                body.conversation_id, body.customer_name, body.order_number, body.order_amount,
                rc, oo, body.reason, bank_name, bank_account, iban,
                body.attachment_name, body.assignee, body.created_by,
                method, ref,
            )
        return row_to_dict(r)


@app.get("/returns/api/requests")
async def list_requests(conversation_id: Optional[int] = None, returns_session: Optional[str] = Cookie(None)):
    p = await pool()
    async with p.acquire() as c:
        if conversation_id is not None:
            # single-conversation lookup (used by the CS tab inside Chatwoot to prefill) — open
            r = await c.fetchrow(
                "SELECT * FROM return_requests WHERE conversation_id=$1 ORDER BY id DESC LIMIT 1",
                conversation_id,
            )
            return row_to_dict(r) if r else JSONResponse(None)
        # full list (bank data) — accountant login required
        if not await current_user(returns_session):
            raise HTTPException(401, "login required")
        rows = await c.fetch("SELECT * FROM return_requests ORDER BY id DESC LIMIT 2000")
        return [row_to_dict(r) for r in rows]


@app.get("/returns/api/requests/{rid}")
async def get_request(rid: int):
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow("SELECT * FROM return_requests WHERE id=$1", rid)
    if not r:
        raise HTTPException(404, "not found")
    return row_to_dict(r)


@app.patch("/returns/api/requests/{rid}/status")
async def set_status(rid: int, body: StatusIn, returns_session: Optional[str] = Cookie(None)):
    user = await current_user(returns_session)
    if not user:
        raise HTTPException(401, "login required")
    if body.status not in ("will", "doing", "done", "rejected", "done_salla"):
        raise HTTPException(400, "invalid status")
    # SERVER-side per-user permission: each role may set ONLY its own statuses.
    # النذير → done_salla only; financial/management → will/doing/done/rejected only.
    if body.status not in allowed_statuses_for(user.get("email")):
        raise HTTPException(403, "غير مصرّح لك بتعيين هذه الحالة")
    if body.status == "rejected" and not (body.reject_reason or "").strip():
        raise HTTPException(400, "سبب الرفض مطلوب عند اختيار مرفوض")
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow("SELECT status, status_history, receipt_path FROM return_requests WHERE id=$1", rid)
        if not r:
            raise HTTPException(404, "not found")
        # A rejected request is FINAL — it can never be reopened or changed.
        # The agent must submit a brand-new return request from Chatwoot.
        if r["status"] == "rejected":
            raise HTTPException(
                409,
                "هذا الطلب مرفوض نهائياً ولا يمكن تغيير حالته. يجب على خدمة العملاء رفع طلب إرجاع جديد.",
            )
        # State-transition guards (root enforcement — not just UI).
        cur = r["status"]
        # A 'done' request is financially final — no status change allowed.
        if cur == "done":
            raise HTTPException(
                409,
                "هذا الطلب منتهٍ مالياً (تم الإرجاع) ولا يمكن تغيير حالته.",
            )
        # Once the financial track has started (will/doing), the accountant may
        # only move it forward to 'done' — not back to will/doing/rejected.
        # (النذير مسار مستقل: done_salla لا يخضع لهذا القيد.)
        if cur in ("will", "doing") and body.status in ACCOUNTANT_STATUSES and body.status != "done":
            raise HTTPException(
                409,
                "بعد بدء المسار المالي لا يمكن إلا التحويل إلى: تم الإرجاع.",
            )
        # 'تم الإرجاع' requires the transfer receipt to be attached first.
        if body.status == "done" and not r["receipt_path"]:
            raise HTTPException(
                400,
                "يجب إرفاق إيصال التحويل (PDF أو صورة) قبل تحويل الطلب إلى: تم الإرجاع.",
            )
        hist = r["status_history"]
        if isinstance(hist, str):
            hist = json.loads(hist)
        hist = list(hist or [])
        entry = {
            "status": body.status,
            "label": STATUS_LABELS[body.status],
            "by": user.get("email") or body.changed_by or "financial@qaydao.com",
            "at": datetime.now(timezone.utc).isoformat(),
        }
        if body.accountant_note is not None and body.accountant_note.strip():
            entry["note"] = body.accountant_note.strip()
        if body.status == "rejected":
            entry["reject_reason"] = body.reject_reason.strip()
        hist.append(entry)
        note_val = body.accountant_note.strip() if (body.accountant_note and body.accountant_note.strip()) else None
        reject_val = body.reject_reason.strip() if body.status == "rejected" else None
        out = await c.fetchrow(
            """UPDATE return_requests
                 SET status=$2, status_history=$3::jsonb,
                     accountant_note=COALESCE($4, accountant_note),
                     reject_reason=CASE WHEN $2='rejected' THEN $5 ELSE reject_reason END
               WHERE id=$1 RETURNING *""",
            rid, body.status, json.dumps(hist), note_val, reject_val,
        )
    return row_to_dict(out)


@app.post("/returns/api/requests/{rid}/contacted")
async def mark_contacted(rid: int):
    # CS-team action: the agent confirms they followed up with the accountant about a
    # rejected request. This ONLY records the follow-up timestamp so the red alert stops
    # counting it. It does NOT change the request status (stays 'rejected') and does NOT
    # touch any accountant/financial logic. Reachable from the team page inside Chatwoot
    # (no accountant login) — same trust model as the read-only team-requests endpoint.
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow("SELECT status, contacted_at FROM return_requests WHERE id=$1", rid)
        if not r:
            raise HTTPException(404, "not found")
        if r["status"] != "rejected":
            raise HTTPException(400, "الإجراء متاح فقط للطلبات المرفوضة")
        out = await c.fetchrow(
            """UPDATE return_requests
                 SET contacted_at = COALESCE(contacted_at, now())
               WHERE id=$1 RETURNING *""",
            rid,
        )
    return row_to_dict(out)


@app.get("/returns/api/me")
async def whoami(returns_session: Optional[str] = Cookie(None)):
    user = await current_user(returns_session)
    if not user:
        raise HTTPException(401, "login required")
    email = (user.get("email") or "").lower()
    return {
        "email": email,
        "display_name": user.get("display_name"),
        # exact set of statuses this user may set — drives which buttons/tabs the UI shows
        "allowed_statuses": allowed_statuses_for(email),
        "is_purchaser": email in PURCHASER_EMAILS,
    }


# ------------------------- Login + Accountant page -------------------------

@app.get("/accountant-login", response_class=HTMLResponse)
async def login_page(returns_session: Optional[str] = Cookie(None)):
    if await current_user(returns_session):
        return RedirectResponse(url="/accountant-returns", status_code=302)
    return HTMLResponse(LOGIN_HTML)


@app.post("/accountant-login")
async def login_submit(email: str = Form(...), password: str = Form(...), remember: Optional[str] = Form(None)):
    email = (email or "").strip().lower()
    p = await pool()
    async with p.acquire() as c:
        u = await c.fetchrow("SELECT email, password_hash FROM accountant_users WHERE email=$1", email)
        if not u or not _check_pw(password, u["password_hash"]):
            return JSONResponse({"ok": False, "error": "البريد الإلكتروني أو كلمة المرور غير صحيحة"}, status_code=401)
        days = SESSION_DAYS_REMEMBER if remember else SESSION_DAYS_DEFAULT
        token = secrets.token_urlsafe(32)
        exp = datetime.now(timezone.utc) + timedelta(days=days)
        await c.execute("INSERT INTO accountant_sessions(token, email, expires_at) VALUES ($1,$2,$3)", token, email, exp)
    resp = JSONResponse({"ok": True, "redirect": "/accountant-returns"})
    resp.set_cookie(SESSION_COOKIE, token, max_age=days * 86400, httponly=True,
                    secure=True, samesite="lax", path="/")
    return resp


@app.get("/accountant-logout")
async def logout(returns_session: Optional[str] = Cookie(None)):
    if returns_session:
        p = await pool()
        async with p.acquire() as c:
            await c.execute("DELETE FROM accountant_sessions WHERE token=$1", returns_session)
    resp = RedirectResponse(url="/accountant-login", status_code=302)
    resp.delete_cookie(SESSION_COOKIE, path="/")
    return resp


@app.get("/accountant-returns", response_class=HTMLResponse)
async def accountant_page(returns_session: Optional[str] = Cookie(None)):
    if not await current_user(returns_session):
        return _require_login_redirect()
    return HTMLResponse(ACCOUNTANT_HTML)


@app.post("/returns/api/requests/{rid}/attachment")
async def upload_attachment(rid: int, file: UploadFile = File(...)):
    mime = (file.content_type or "").split(";")[0].strip().lower()
    if mime not in ALLOWED_MIME:
        raise HTTPException(400, "نوع ملف غير مسموح. المسموح: PDF, JPG, PNG, WEBP")
    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, "حجم الملف يتجاوز 10 ميجابايت")
    if not data:
        raise HTTPException(400, "الملف فارغ")

    p = await pool()
    async with p.acquire() as c:
        exists = await c.fetchrow("SELECT attachment_path FROM return_requests WHERE id=$1", rid)
        if not exists:
            raise HTTPException(404, "not found")
        # remove previous file if present
        old = exists["attachment_path"]
        if old:
            try:
                Path(old).unlink(missing_ok=True)
            except Exception:
                pass
        ext = ALLOWED_MIME[mime]
        fname = f"{rid}_{uuid.uuid4().hex}{ext}"
        fpath = UPLOAD_DIR / fname
        fpath.write_bytes(data)
        orig = os.path.basename(file.filename or f"attachment{ext}")
        r = await c.fetchrow(
            """UPDATE return_requests
                 SET attachment_name=$2, attachment_path=$3, attachment_mime=$4
               WHERE id=$1 RETURNING *""",
            rid, orig, str(fpath), mime,
        )
    return row_to_dict(r)


@app.get("/returns/api/requests/{rid}/attachment")
async def download_attachment(rid: int, returns_session: Optional[str] = Cookie(None)):
    if not await current_user(returns_session):
        raise HTTPException(401, "login required")
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow(
            "SELECT attachment_name, attachment_path, attachment_mime FROM return_requests WHERE id=$1",
            rid,
        )
    if not r or not r["attachment_path"]:
        raise HTTPException(404, "لا يوجد ملف مرفق")
    fp = Path(r["attachment_path"])
    if not fp.exists():
        raise HTTPException(404, "الملف غير موجود على الخادم")
    return FileResponse(
        str(fp),
        media_type=r["attachment_mime"] or "application/octet-stream",
        filename=r["attachment_name"] or fp.name,
    )


@app.post("/returns/api/requests/{rid}/receipt")
async def upload_receipt(rid: int, file: UploadFile = File(...), returns_session: Optional[str] = Cookie(None)):
    # Accountant-only: transfer receipt proving the refund was made.
    if not await current_user(returns_session):
        raise HTTPException(401, "login required")
    mime = (file.content_type or "").split(";")[0].strip().lower()
    if mime not in ALLOWED_MIME:
        raise HTTPException(400, "نوع ملف غير مسموح. المسموح: PDF, JPG, PNG, WEBP")
    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, "حجم الملف يتجاوز 10 ميجابايت")
    if not data:
        raise HTTPException(400, "الملف فارغ")

    p = await pool()
    async with p.acquire() as c:
        row = await c.fetchrow("SELECT receipt_path, status FROM return_requests WHERE id=$1", rid)
        if not row:
            raise HTTPException(404, "not found")
        if row["status"] == "rejected":
            raise HTTPException(409, "هذا الطلب مرفوض نهائياً ولا يمكن إرفاق إيصال له.")
        if row["receipt_path"]:
            try:
                Path(row["receipt_path"]).unlink(missing_ok=True)
            except Exception:
                pass
        ext = ALLOWED_MIME[mime]
        fname = f"receipt_{rid}_{uuid.uuid4().hex}{ext}"
        fpath = UPLOAD_DIR / fname
        fpath.write_bytes(data)
        orig = os.path.basename(file.filename or f"receipt{ext}")
        r = await c.fetchrow(
            """UPDATE return_requests
                 SET receipt_name=$2, receipt_path=$3, receipt_mime=$4
               WHERE id=$1 RETURNING *""",
            rid, orig, str(fpath), mime,
        )
    return row_to_dict(r)


@app.get("/returns/api/requests/{rid}/receipt")
async def download_receipt(rid: int):
    # Open: the agent needs it (from the team page) to reply to the customer with proof of transfer.
    p = await pool()
    async with p.acquire() as c:
        r = await c.fetchrow(
            "SELECT receipt_name, receipt_path, receipt_mime FROM return_requests WHERE id=$1", rid
        )
    if not r or not r["receipt_path"]:
        raise HTTPException(404, "لا يوجد إيصال مرفق")
    fp = Path(r["receipt_path"])
    if not fp.exists():
        raise HTTPException(404, "الملف غير موجود على الخادم")
    return FileResponse(
        str(fp),
        media_type=r["receipt_mime"] or "application/octet-stream",
        filename=r["receipt_name"] or fp.name,
    )


@app.get("/returns/api/config")
async def config():
    return {"reasons": REASONS, "assignees": ASSIGNEES, "status_labels": STATUS_LABELS}


LOGIN_HTML = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>تسجيل الدخول — المرجعات — QAYDAO</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{--brand:#1f5f5b;--brandink:#12403d;--ink:#1f2b3a;--soft:#5a6b7d;--line:#e4e9ee;--bg:#eef2f4}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Cairo,system-ui,sans-serif;background:linear-gradient(135deg,#eef2f4,#e4efee);min-height:100vh;display:grid;place-items:center;padding:20px}
.card{background:#fff;border:1px solid var(--line);border-radius:20px;box-shadow:0 10px 40px rgba(31,43,58,.12);width:100%;max-width:400px;overflow:hidden}
.top{background:var(--brand);color:#fff;padding:28px 26px;text-align:center}
.logo{width:52px;height:52px;border-radius:14px;background:rgba(255,255,255,.16);display:grid;place-items:center;font-weight:800;font-size:24px;margin:0 auto 12px}
.top h1{font-size:19px;font-weight:800}
.top p{font-size:12.5px;opacity:.9;margin-top:4px}
.body{padding:26px 26px 28px}
.f{margin-bottom:16px}
label{display:block;font-size:13px;font-weight:600;color:var(--soft);margin-bottom:6px}
input[type=email],input[type=password]{width:100%;font-family:inherit;font-size:14.5px;color:var(--ink);background:#f8fafb;border:1px solid var(--line);border-radius:11px;padding:12px 14px;transition:.15s}
input:focus{outline:none;border-color:var(--brand);background:#fff}
.remember{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--soft);cursor:pointer;margin-bottom:18px}
.remember input{width:17px;height:17px;accent-color:var(--brand);cursor:pointer}
.btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:12px;padding:13px;font-family:inherit;font-size:15px;font-weight:700;cursor:pointer;transition:.15s}
.btn:hover{background:var(--brandink)}
.btn:disabled{opacity:.6;cursor:default}
.err{background:#fdeded;color:#c0392b;border:1px solid #f3c6c1;border-radius:10px;padding:10px 13px;font-size:13px;font-weight:600;margin-bottom:16px;display:none}
.err.show{display:block}
.foot{text-align:center;font-size:11.5px;color:var(--soft);margin-top:18px}
</style></head><body>
<div class="card">
  <div class="top">
    <div class="logo">Q</div>
    <h1>إدارة المرجعات</h1>
    <p>تسجيل دخول المحاسبة والإدارة</p>
  </div>
  <div class="body">
    <div class="err" id="err"></div>
    <div class="f"><label>البريد الإلكتروني</label>
      <input type="email" id="email" dir="ltr" style="text-align:right" placeholder="financial@qaydao.com" autocomplete="username"></div>
    <div class="f"><label>كلمة المرور</label>
      <input type="password" id="password" placeholder="••••••••" autocomplete="current-password"></div>
    <label class="remember"><input type="checkbox" id="remember"> تذكّرني على هذا الجهاز</label>
    <button class="btn" id="btn" onclick="doLogin()">تسجيل الدخول</button>
    <div class="foot">QAYDAO · دخول آمن — البيانات البنكية محمية</div>
  </div>
</div>
<script>
var e=document.getElementById("email"),p=document.getElementById("password"),
    r=document.getElementById("remember"),b=document.getElementById("btn"),er=document.getElementById("err");
try{var saved=localStorage.getItem("qd_ret_email");if(saved){e.value=saved;r.checked=true;p.focus()}else{e.focus()}}catch(x){}
function showErr(m){er.textContent=m;er.classList.add("show")}
function doLogin(){
  er.classList.remove("show");
  var em=e.value.trim(),pw=p.value;
  if(!em||!pw){showErr("يرجى إدخال البريد وكلمة المرور");return}
  b.disabled=true;b.textContent="جارٍ الدخول…";
  var fd=new URLSearchParams();fd.append("email",em);fd.append("password",pw);if(r.checked)fd.append("remember","1");
  fetch("/accountant-login",{method:"POST",headers:{"Content-Type":"application/x-www-form-urlencoded"},body:fd.toString(),credentials:"same-origin"})
    .then(function(res){return res.json().then(function(d){return {ok:res.ok,d:d}})})
    .then(function(o){
      if(o.ok&&o.d.ok){
        try{if(r.checked)localStorage.setItem("qd_ret_email",em);else localStorage.removeItem("qd_ret_email")}catch(x){}
        location.href=o.d.redirect||"/accountant-returns";
      }else{showErr((o.d&&o.d.error)||"تعذّر تسجيل الدخول");b.disabled=false;b.textContent="تسجيل الدخول"}
    })
    .catch(function(){showErr("خطأ في الاتصال، حاول مجدداً");b.disabled=false;b.textContent="تسجيل الدخول"});
}
p.addEventListener("keydown",function(ev){if(ev.key==="Enter")doLogin()});
e.addEventListener("keydown",function(ev){if(ev.key==="Enter")p.focus()});
</script></body></html>"""


ACCOUNTANT_HTML = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>صفحة المحاسب — المرجعات — QAYDAO</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{--bg:#f4f6f8;--surface:#fff;--ink:#1f2b3a;--soft:#5a6b7d;--line:#e4e9ee;--brand:#1f5f5b;--brandsoft:#e8f1f0;--brandink:#12403d;--accent:#c8892a;--ok:#1f7a4d;--oksoft:#e6f4ec;--amber:#b5791d;--ambersoft:#fbf0d9;--info:#2c5a86;--infosoft:#eef4fb;--infoline:#cfe0f2}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Cairo,system-ui,sans-serif;background:var(--bg);color:var(--ink);line-height:1.7;padding-bottom:50px}
.ltr{direction:ltr;unicode-bidi:isolate;display:inline-block}
.wrap{max-width:1200px;margin:0 auto;padding:0 20px}
.topbar{background:repeating-linear-gradient(-45deg,#12403d,#12403d 12px,#0f3a37 12px,#0f3a37 24px);color:#fff;text-align:center;padding:8px;font-size:12.5px;font-weight:600}
.topbar b{color:#ffd479}
header{background:var(--surface);border-bottom:1px solid var(--line);padding:20px 0}
.hrow{display:flex;align-items:center;gap:14px;flex-wrap:wrap}
.logo{width:44px;height:44px;border-radius:11px;background:var(--brand);color:#fff;display:grid;place-items:center;font-weight:800;font-size:19px}
h1{font-size:20px;font-weight:800}
header p{font-size:13px;color:var(--soft);margin-top:2px}
.motiv-acc{margin-inline-start:auto;max-width:280px;font-size:12.5px;font-weight:600;color:var(--brand);background:linear-gradient(135deg,var(--brandsoft),#f3f9f8);border:1px solid #d5e6e4;border-radius:12px;padding:10px 14px;line-height:1.5}
.logout-btn{font-family:inherit;font-size:12.5px;font-weight:700;color:#c0392b;background:#fdeded;border:1px solid #f3c6c1;border-radius:10px;padding:8px 14px;text-decoration:none;white-space:nowrap}
.logout-btn:hover{background:#fbdddd}
.pill{margin-inline-start:auto;background:var(--oksoft);color:var(--ok);border:1px solid #bfe3cd;border-radius:999px;padding:6px 13px;font-size:12px;font-weight:700}
.tools{display:flex;gap:10px;align-items:center;margin:18px 0;flex-wrap:wrap}
.tools select,.tools input{font-family:inherit;font-size:13.5px;padding:9px 12px;border:1px solid var(--line);border-radius:10px;background:#fff}
.tabs{display:flex;gap:8px;flex-wrap:wrap;margin:18px 0 6px}
.tab{font-family:inherit;font-size:13px;font-weight:700;cursor:pointer;border:1px solid var(--line);background:#fff;color:var(--soft);border-radius:999px;padding:8px 16px;transition:.15s;display:flex;align-items:center;gap:7px}
.tab:hover{border-color:var(--brand);color:var(--brand)}
.tab.active{background:var(--brand);color:#fff;border-color:var(--brand)}
.tab .cnt{font-size:11px;font-weight:700;background:rgba(0,0,0,.12);border-radius:999px;padding:1px 7px;min-width:18px;text-align:center}
.tab.active .cnt{background:rgba(255,255,255,.28)}
.tab-old{border-style:dashed;color:#8a92a0}
.tab-old:hover{border-color:#8a92a0;color:#5a6b7d}
.tab-old.active{background:#5a6b7d;border-color:#5a6b7d;border-style:solid;color:#fff}
.refresh{margin-inline-start:auto;background:var(--brand);color:#fff;border:none;border-radius:10px;padding:9px 16px;font-family:inherit;font-weight:700;font-size:13px;cursor:pointer}
.count{font-size:12.5px;color:var(--soft)}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:16px}
.card{background:var(--surface);border:1px solid var(--line);border-radius:14px;box-shadow:0 1px 2px rgba(31,43,58,.04),0 6px 20px rgba(31,43,58,.05);overflow:hidden}
.chead{padding:14px 16px;border-bottom:1px solid var(--line);display:flex;align-items:center;gap:9px}
.chead .nm{font-weight:700;font-size:15px}
.chead .st{margin-inline-start:auto;font-size:11px;font-weight:700;padding:4px 10px;border-radius:999px}
.st.new{background:#eef1f4;color:#5a6b7d}
.st.will{background:var(--ambersoft);color:var(--amber)}
.st.doing{background:var(--infosoft);color:var(--info)}
.st.done{background:var(--oksoft);color:var(--ok)}
.st.rejected{background:#fdeded;color:#c0392b}
.cbody{padding:14px 16px}
.rowf{display:flex;justify-content:space-between;gap:10px;padding:5px 0;font-size:13px;border-bottom:1px dashed #eef1f4}
.rowf:last-of-type{border-bottom:none}
.rowf .k{color:var(--soft);font-weight:500;flex:0 0 auto}
.rowf .v{font-weight:600;text-align:end}
.copyv{display:inline-flex;gap:6px;align-items:center}
.cbtn{background:var(--brandsoft);color:var(--brandink);border:1px solid #cfe3e1;border-radius:7px;padding:2px 8px;font-family:inherit;font-size:11px;font-weight:700;cursor:pointer}
.cbtn.ok{background:var(--oksoft);color:var(--ok);border-color:#bfe3cd}
.olink{color:var(--brand);font-weight:700;text-decoration:none;border-bottom:1px dashed var(--brand)}
.sbtns{display:grid;grid-template-columns:repeat(2,1fr);gap:7px;margin-top:12px}
.sbtn{font-family:inherit;font-size:12px;font-weight:700;cursor:pointer;border-radius:9px;padding:9px 4px;border:1.5px solid transparent;transition:.15s}
.sbtn.will{background:var(--ambersoft);color:var(--amber);border-color:#eddcae}
.sbtn.doing{background:var(--infosoft);color:var(--info);border-color:var(--infoline)}
.sbtn.done{background:var(--oksoft);color:var(--ok);border-color:#c5dddb}
.sbtn.rejected{background:#fdeded;color:#c0392b;border-color:#f3c6c1}
.sbtn.done_salla{background:#eef4ff;color:#3b5bdb;border-color:#c7d6f5}
.st.done_salla{background:#eef4ff;color:#3b5bdb}
/* شريط "طلب منتج جديد" — عرض بصري فقط في صفحة المحاسب (لا يمس أي حالة/صلاحية/قاعدة) */
.card.is-new{position:relative}
.ribbon-new{display:flex;align-items:center;gap:6px;background:linear-gradient(90deg,var(--ok),#2f9e6f);color:#fff;font-weight:800;font-size:12px;letter-spacing:.2px;padding:7px 14px}
.ribbon-new::before{content:"\1F195"}
.sbtn:hover{transform:translateY(-1px)}
.sbtn.active{box-shadow:0 0 0 3px rgba(31,95,91,.14)}
.qd-note-in{width:100%;font-family:inherit;font-size:12.5px;color:#1f2b3a;background:#f8fafb;border:1px solid var(--line);border-radius:9px;padding:8px 10px;margin-top:10px;resize:vertical}
.qd-note-in:focus{outline:none;border-color:var(--brand);background:#fff}
.qd-note-lbl{font-size:11.5px;font-weight:600;color:var(--soft);margin-top:11px;margin-bottom:2px}
.qd-reject-wrap{display:none;margin-top:8px}
.qd-reject-wrap.show{display:block}
.qd-reject-wrap textarea{border-color:#f3c6c1;background:#fdeded}
.rcpt-wrap{display:none;margin-top:12px;background:var(--oksoft);border:1px solid #bfe3cd;border-radius:11px;padding:11px 13px}
.rcpt-wrap.show{display:block}
.rcpt-head{font-size:12.5px;font-weight:700;color:var(--ok);margin-bottom:8px}
.rcpt-opt{font-weight:500;color:var(--soft);font-size:11px}
.rcpt-btn{display:inline-block;font-family:inherit;font-size:12px;font-weight:700;cursor:pointer;border-radius:8px;padding:7px 13px;background:#fff;color:var(--ok);border:1px solid #bfe3cd;transition:.15s}
.rcpt-btn:hover{background:#d9f0e2}
.rcpt-up{width:100%;text-align:center}
.rcpt-has{display:flex;align-items:center;gap:9px;flex-wrap:wrap}
.rcpt-has .olink{flex:1;min-width:0;word-break:break-all}
.rcpt-msg{font-size:11.5px;font-weight:600;margin-top:7px}
.rcpt-msg.ok{color:var(--ok)}.rcpt-msg.err{color:#c0392b}
.rcpt-req{font-weight:700;color:#c0392b;font-size:11px}
.sbtn.needs-rcpt{opacity:.55}
.locked{margin-top:12px;background:#fdeded;border:1px solid #f3c6c1;border-radius:11px;padding:12px 14px;font-size:12.5px;font-weight:600;color:#c0392b;line-height:1.6}
.done-final{margin-top:12px;background:var(--oksoft);border:1px solid #b7e4cd;border-radius:11px;padding:12px 14px;font-size:12.5px;font-weight:600;color:var(--ok);line-height:1.6}
.locked span{font-weight:500;color:#8a4038;font-size:11.5px}
.rejsend{flex:1;font-family:inherit;font-size:12.5px;font-weight:700;cursor:pointer;border-radius:9px;padding:9px 6px;border:none;background:#c0392b;color:#fff;transition:.15s}
.rejsend:hover{background:#a5301f}
.rejcancel{font-family:inherit;font-size:12.5px;font-weight:600;cursor:pointer;border-radius:9px;padding:9px 14px;border:1px solid var(--line);background:#f8fafb;color:var(--soft)}
.pick-lbl{font-size:11.5px;font-weight:600;color:var(--soft);margin:14px 0 8px;text-align:center}
.sendbtn{width:100%;margin-top:11px;font-family:inherit;font-size:13.5px;font-weight:800;cursor:pointer;border:none;border-radius:11px;padding:12px;background:var(--brand);color:#fff;transition:.15s}
.sendbtn:hover:not(:disabled){background:var(--brandink)}
.sendbtn:disabled{background:#e4e9ee;color:#a3adb8;cursor:not-allowed}
.hist{margin-top:11px;font-size:11.5px;color:var(--soft);background:#f8fafb;border-radius:9px;padding:9px 11px;display:none}
.hist.show{display:block}
.hist b{color:var(--ink)}
.empty{text-align:center;color:var(--soft);padding:60px 20px;font-size:15px}
.toast{position:fixed;bottom:20px;inset-inline-start:20px;background:var(--ink);color:#fff;padding:12px 18px;border-radius:11px;font-size:13.5px;font-weight:600;opacity:0;transform:translateY(10px);transition:.25s;z-index:99;pointer-events:none}
.toast.show{opacity:1;transform:none}
footer{text-align:center;margin-top:26px;font-size:11.5px;color:var(--soft)}
</style></head><body>
<div class="topbar">💼 صفحة المحاسب — <b>قراءة فقط</b> — يُسمح فقط بتغيير حالة الإرجاع</div>
<header><div class="wrap hrow">
  <div class="logo">Q</div>
  <div><h1>إدارة المرجعات — المحاسبة</h1><p>عرض طلبات الإرجاع الواردة من خدمة العملاء وتحديث حالتها.</p></div>
  <div class="motiv-acc">✨ دقّتك في المراجعة تحمي حقوق العملاء والمتجر — شكراً لك</div>
  <span class="pill" id="live">● متصل</span>
  <a href="/accountant-logout" class="logout-btn">خروج</a>
</div></header>
<div class="wrap">
  <div class="tabs" id="tabs">
    <button class="tab active" data-tab="new" onclick="setTab('new',this)">جديدة <span class="cnt" id="cnt-new">0</span></button>
    <button class="tab" data-tab="will" onclick="setTab('will',this)">سيتم الإرجاع <span class="cnt" id="cnt-will">0</span></button>
    <button class="tab" data-tab="doing" onclick="setTab('doing',this)">جاري الإرجاع <span class="cnt" id="cnt-doing">0</span></button>
    <button class="tab" data-tab="done" onclick="setTab('done',this)">تم الإرجاع <span class="cnt" id="cnt-done">0</span></button>
    <button class="tab" data-tab="rejected" onclick="setTab('rejected',this)">مرفوض <span class="cnt" id="cnt-rejected">0</span></button>
    <button class="tab" data-tab="done_salla" id="tab-done_salla" style="display:none" onclick="setTab('done_salla',this)">تم الإرجاع في سلة <span class="cnt" id="cnt-done_salla">0</span></button>
    <button class="tab tab-old" data-tab="old_done" onclick="setTab('old_done',this)">قديمة — تم الإرجاع <span class="cnt" id="cnt-old_done">0</span></button>
    <button class="tab tab-old" data-tab="old_rejected" onclick="setTab('old_rejected',this)">قديمة — مرفوضة <span class="cnt" id="cnt-old_rejected">0</span></button>
    <button class="tab" data-tab="all" onclick="setTab('all',this)">الكل <span class="cnt" id="cnt-all">0</span></button>
  </div>
  <div class="tools">
    <input id="fsearch" placeholder="بحث بالاسم أو رقم الطلب أو المحادثة…" oninput="render()">
    <span class="count" id="count"></span>
    <button class="refresh" onclick="load()">تحديث ⟳</button>
  </div>
  <div class="grid" id="grid"></div>
  <div class="empty" id="empty" style="display:none">لا توجد طلبات في هذا القسم.</div>
  <footer>QAYDAO · صفحة المحاسب · التخزين في خدمة مستقلة — لا يوجد ربط فعلي مع قاعدة بيانات Chatwoot أو سلة</footer>
</div>
<div class="toast" id="toast"></div>
<script>
var API="/returns/api/requests";
var DATA=[];
var SL={new:"جديد",will:"سيتم الإرجاع",doing:"جاري الإرجاع",done:"تم الإرجاع",rejected:"مرفوض",done_salla:"تم الإرجاع في سلة"};
var ME={email:"",allowed_statuses:["will","doing","done","rejected"],is_purchaser:false};
function canStatus(st){return ME.allowed_statuses&&ME.allowed_statuses.indexOf(st)>=0;}
function canSalla(){return canStatus("done_salla");}
function esc(s){return (s==null?"":String(s)).replace(/[&<>"]/g,function(c){return{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]})}
function toast(m){var t=document.getElementById("toast");t.textContent=m;t.classList.add("show");setTimeout(function(){t.classList.remove("show")},2200)}
function copy(v,btn){navigator.clipboard.writeText(v).then(function(){var o=btn.textContent;btn.textContent="تم ✓";btn.classList.add("ok");setTimeout(function(){btn.textContent=o;btn.classList.remove("ok")},1400)})}
function load(){fetch(API,{credentials:"same-origin"}).then(function(r){if(r.status===401){location.href="/accountant-login";return null}return r.json()}).then(function(d){if(d===null)return;DATA=Array.isArray(d)?d:[];render()}).catch(function(){document.getElementById("live").textContent="● غير متصل";document.getElementById("live").style.color="#c0392b"})}
var CURTAB="new";
var OLD_DAYS=7;  // "قديمة" = مضى أكثر من 7 أيام على إغلاق الطلب (آخر تغيير حالة)
function closedAt(x){
  // use the timestamp of the LAST status change (status_history), falling back to updated_at
  var h=x.status_history;
  if(h&&h.length){
    var last=h[h.length-1];
    if(last&&last.at)return last.at;
  }
  return x.updated_at||x.created_at;
}
function isOld(x){
  var t=closedAt(x);
  if(!t)return false;
  var d=new Date(t);
  if(isNaN(d))return false;
  return (Date.now()-d.getTime()) > OLD_DAYS*86400000;
}
function setTab(t,btn){
  CURTAB=t;
  var tabs=document.querySelectorAll("#tabs .tab");
  for(var i=0;i<tabs.length;i++)tabs[i].classList.remove("active");
  if(btn)btn.classList.add("active");
  render();
}
function inTab(x){
  switch(CURTAB){
    case "all":          return true;
    case "new":          return x.status==="new";
    case "will":         return x.status==="will";
    case "doing":        return x.status==="doing";
    case "done":         return x.status==="done"     && !isOld(x);
    case "rejected":     return x.status==="rejected" && !isOld(x);
    case "done_salla":   return x.status==="done_salla";
    case "old_done":     return x.status==="done"     &&  isOld(x);
    case "old_rejected": return x.status==="rejected" &&  isOld(x);
    default:             return true;
  }
}
function updateCounts(){
  var c={new:0,will:0,doing:0,done:0,rejected:0,done_salla:0,old_done:0,old_rejected:0};
  DATA.forEach(function(x){
    var old=isOld(x);
    if(x.status==="new")c.new++;
    else if(x.status==="will")c.will++;
    else if(x.status==="doing")c.doing++;
    else if(x.status==="done"){ old?c.old_done++:c.done++; }
    else if(x.status==="rejected"){ old?c.old_rejected++:c.rejected++; }
    else if(x.status==="done_salla")c.done_salla++;
  });
  var set=function(id,n){var e=document.getElementById(id);if(e)e.textContent=n};
  set("cnt-new",c.new);set("cnt-will",c.will);set("cnt-doing",c.doing);
  set("cnt-done",c.done);set("cnt-rejected",c.rejected);set("cnt-done_salla",c.done_salla);
  set("cnt-old_done",c.old_done);set("cnt-old_rejected",c.old_rejected);
  set("cnt-all",DATA.length);
}
function render(){
  updateCounts();
  var q=document.getElementById("fsearch").value.trim().toLowerCase();
  var list=DATA.filter(function(x){
    if(!inTab(x))return false;
    if(q){var h=((x.customer_name||"")+" "+(x.order_number||"")+" "+(x.conversation_id||"")).toLowerCase();if(h.indexOf(q)<0)return false}
    return true;
  });
  document.getElementById("count").textContent=list.length+" طلب";
  var g=document.getElementById("grid"),e=document.getElementById("empty");
  if(!list.length){g.innerHTML="";e.style.display="block";return}
  e.style.display="none";
  g.innerHTML=list.map(card).join("");
}
// طلب حديث = تاريخ طلب المنتجات الأصلي خلال 3 أيام أو أقل من اليوم (مقارنة تواريخ فقط)
function isRecentOrder(d){
  if(!d)return false;
  var m=/^(\d{4})-(\d{2})-(\d{2})/.exec(String(d));
  if(!m)return false;
  var oo=new Date(+m[1],+m[2]-1,+m[3]);
  var now=new Date();var today=new Date(now.getFullYear(),now.getMonth(),now.getDate());
  var diff=Math.floor((today-oo)/86400000);
  return diff>=0 && diff<=3;
}
// أزرار الحالة المسموح عرضها للمستخدم الحالي حسب حالة الطلب (يطابق قيود السيرفر).
// المحاسب: new/done_salla => الأربعة | will/doing => done فقط | done/rejected => لا شيء.
// النذير: done_salla => يظهر فقط على new/will/doing.
function visibleActions(x){
  var out=[];
  var s=x.status;
  // حالات المحاسب المالية
  if(s==="new" || s==="done_salla"){
    ["will","doing","done","rejected"].forEach(function(st){ if(canStatus(st))out.push(st); });
  } else if(s==="will" || s==="doing"){
    if(canStatus("done"))out.push("done");
  }
  // زر النذير — مستقل، فقط على new/will/doing
  if(canStatus("done_salla") && (s==="new"||s==="will"||s==="doing")){
    out.push("done_salla");
  }
  return out;
}
// صفوف طريقة الاسترداد: تابي/تمارا/مدفوع => رقم/مرجع واحد | الحساب البنكي => البنك+الحساب+الآيبان.
// السجلات القديمة بلا refund_method تُعامل كحساب بنكي (سلوك ما قبل الميزة).
var RM={tabby:"تابي",tamara:"تمارا",madfooat:"مدفوع",bank_account:"الحساب البنكي"};
function refundRows(x){
  var m=x.refund_method||"bank_account";
  var head='<div class="rowf"><span class="k">طريقة الاسترداد</span><span class="v"><b>'+esc(RM[m]||m)+'</b></span></div>';
  if(m==="bank_account"){
    var acc=esc(x.bank_account||"—"),iban=esc(x.iban||"—");
    return head+
      '<div class="rowf"><span class="k">البنك</span><span class="v">'+esc(x.bank_name||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">الحساب البنكي</span><span class="v copyv"><span class="ltr">'+acc+'</span><button class="cbtn" onclick="copy(\''+acc+'\',this)">نسخ</button></span></div>'+
      '<div class="rowf"><span class="k">الآيبان</span><span class="v copyv"><span class="ltr">'+iban+'</span><button class="cbtn" onclick="copy(\''+iban+'\',this)">نسخ</button></span></div>';
  }
  var ref=esc(x.refund_reference||"—");
  return head+
    '<div class="rowf"><span class="k">رقم / مرجع '+esc(RM[m]||m)+'</span>'+
    '<span class="v copyv"><span class="ltr">'+ref+'</span>'+
    '<button class="cbtn" onclick="copy(\''+ref+'\',this)">نسخ</button></span></div>';
}
function card(x){
  var isNew=isRecentOrder(x.original_order_at);
  var order=x.order_number?('<a class="olink ltr" href="#" onclick="return orderClick(\''+esc(x.order_number)+'\')">'+esc(x.order_number)+'</a>'):"—";
  var histRows=(x.status_history||[]).map(function(h){return '<div><b>'+esc(h.label)+'</b> · '+esc((h.at||"").replace("T"," ").slice(0,16))+' · '+esc(h.by||"")+'</div>'}).join("");
  return '<div class="card'+(isNew?' is-new':'')+'">'+
    (isNew?'<div class="ribbon-new">طلب منتج جديد</div>':'')+
    '<div class="chead"><span class="nm">'+esc(x.customer_name||"—")+'</span><span class="st '+x.status+'">'+esc(SL[x.status]||x.status)+'</span></div>'+
    '<div class="cbody">'+
      '<div class="rowf"><span class="k">اسم العميل</span><span class="v">'+esc(x.customer_name||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">رقم الطلب</span><span class="v">'+order+'</span></div>'+
      '<div class="rowf"><span class="k">مبلغ الطلب</span><span class="v">'+esc(x.order_amount||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">سبب الإرجاع</span><span class="v">'+esc(x.reason||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">تاريخ طلب الإرجاع</span><span class="v ltr">'+esc(x.return_created_at||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">تاريخ الطلب الأصلي</span><span class="v ltr">'+esc(x.original_order_at||"—")+'</span></div>'+
      refundRows(x)+
      '<div class="rowf"><span class="k">الموظف المسؤول</span><span class="v">'+esc(x.assignee||"—")+'</span></div>'+
      '<div class="rowf"><span class="k">ملف الحساب البنكي</span><span class="v">'+
        (x.attachment_name?('<a class="olink" href="/returns/api/requests/'+x.id+'/attachment" target="_blank">\u2B07 '+esc(x.attachment_name)+'</a>'):'—')+
      '</span></div>'+
      (x.reject_reason?('<div class="rowf"><span class="k" style="color:#c0392b">سبب الرفض</span><span class="v" style="color:#c0392b">'+esc(x.reject_reason)+'</span></div>'):'')+
      (x.status==="rejected"
        ? ('<div class="locked">\uD83D\uDD12 هذا الطلب <b>مرفوض نهائياً</b> ولا يمكن تغيير حالته.<br>'+
           '<span>لإعادة فتحه يجب على خدمة العملاء رفع طلب إرجاع جديد من الشات.</span></div>')
      : x.status==="done"
        ? ('<div class="done-final">\u2705 هذا الطلب <b>منتهٍ مالياً (تم الإرجاع)</b> — للعرض فقط، لا يمكن تغيير حالته.</div>')
        : (function(){
            var acts=visibleActions(x);
            if(!acts.length) return '';
            var LBL={will:"سيتم الإرجاع",doing:"جاري الإرجاع",done:"تم الإرجاع",rejected:"مرفوض",done_salla:"تم الإرجاع في سلة"};
            var btns=acts.map(function(st){
              return '<button class="sbtn '+st+'" data-st="'+st+'" onclick="pick('+x.id+',\''+st+'\',this)">'+LBL[st]+'</button>';
            }).join("");
            return '<div class="pick-lbl">اختر الحالة الجديدة ثم اضغط إرسال</div>'+
          '<div class="sbtns" id="sb_'+x.id+'">'+
            btns+
          '</div>'+
          // receipt box — hidden until "تم الإرجاع" is picked
          '<div class="rcpt-wrap" id="rcptw_'+x.id+'">'+
            '<div class="rcpt-head">🧾 إيصال التحويل <span class="rcpt-req">(إلزامي)</span></div>'+
            (x.receipt_name
              ? ('<div class="rcpt-has"><a class="olink" href="/returns/api/requests/'+x.id+'/receipt" target="_blank">\u2B07 '+esc(x.receipt_name)+'</a>'+
                 '<label class="rcpt-btn">استبدال<input type="file" accept=".pdf,.jpg,.jpeg,.png,.webp,application/pdf,image/*" style="display:none" onchange="uploadReceipt('+x.id+',this)"></label></div>')
              : ('<label class="rcpt-btn rcpt-up">\u2B06 رفع إيصال (PDF أو صورة)<input type="file" accept=".pdf,.jpg,.jpeg,.png,.webp,application/pdf,image/*" style="display:none" onchange="uploadReceipt('+x.id+',this)"></label>'))+
            '<div class="rcpt-msg" id="rcptm_'+x.id+'"></div>'+
          '</div>'+
          // reject reason box — hidden until "مرفوض" is picked
          '<div class="qd-reject-wrap" id="rejw_'+x.id+'">'+
            '<div class="qd-note-lbl" style="color:#c0392b">سبب الرفض (إلزامي)</div>'+
            '<textarea class="qd-note-in" id="rej_'+x.id+'" rows="2" placeholder="اكتب سبب الرفض…"></textarea>'+
          '</div>'+
          '<button class="sendbtn" id="send_'+x.id+'" disabled onclick="submitStatus('+x.id+',this)">إرسال</button>';
          })())+
      (histRows?'<div class="hist show">'+histRows+'</div>':'')+
    '</div></div>';
}
function orderClick(n){toast("رقم الطلب "+n+" — الربط مع سلة سيُفعّل لاحقاً.");return false}

/* --- selection model: pick a status, fill what it needs, then press إرسال --- */
var PICKED={};   // {requestId: "will"|"doing"|"done"|"rejected"}

function pick(id,st,btn){
  PICKED[id]=st;
  // highlight the chosen button only
  var box=document.getElementById("sb_"+id);
  if(box){
    var bs=box.querySelectorAll(".sbtn");
    for(var i=0;i<bs.length;i++)bs[i].classList.remove("active");
  }
  if(btn)btn.classList.add("active");
  // show only the box the chosen status needs
  var rc=document.getElementById("rcptw_"+id);
  var rj=document.getElementById("rejw_"+id);
  if(rc)rc.classList.toggle("show",st==="done");
  if(rj)rj.classList.toggle("show",st==="rejected");
  if(st==="rejected"){var t=document.getElementById("rej_"+id);if(t)t.focus()}
  var send=document.getElementById("send_"+id);
  if(send){send.disabled=false;send.textContent="إرسال — "+SL[st]}
}

function uploadReceipt(id,input){
  var f=(input&&input.files&&input.files.length)?input.files[0]:null;
  var m=document.getElementById("rcptm_"+id);
  if(!f)return;
  if(f.size>10*1024*1024){m.className="rcpt-msg err";m.textContent="حجم الملف يتجاوز 10 ميجابايت.";input.value="";return}
  m.className="rcpt-msg";m.textContent="جارٍ رفع الإيصال…";
  var fd=new FormData();fd.append("file",f);
  fetch(API+"/"+id+"/receipt",{method:"POST",credentials:"same-origin",body:fd})
    .then(function(r){if(!r.ok)return r.json().then(function(e){throw {msg:(e&&e.detail)||"تعذّر رفع الإيصال"}});return r.json()})
    .then(function(u){
      var i=DATA.findIndex(function(d){return d.id===id});if(i>=0)DATA[i]=u;
      m.className="rcpt-msg ok";m.textContent="تم رفع الإيصال ✓ اضغط إرسال لإتمام الإرجاع.";
      toast("تم رفع إيصال التحويل ✓ اضغط إرسال.");
      // keep the current selection alive after re-render
      var keep=PICKED[id];
      render();
      if(keep){
        var b=document.querySelector('#sb_'+id+' .sbtn[data-st="'+keep+'"]');
        if(b)pick(id,keep,b);
      }
    })
    .catch(function(e){m.className="rcpt-msg err";m.textContent=(e&&e.msg)||"تعذّر رفع الإيصال.";input.value=""});
}

function submitStatus(id,btn){
  var st=PICKED[id];
  if(!st){toast("اختر الحالة أولاً.");return}
  var payload={status:st,changed_by:"financial@qaydao.com"};

  if(st==="done"){
    var rec=DATA.find(function(d){return d.id===id});
    if(!rec||!rec.receipt_name){
      toast("يجب رفع إيصال التحويل قبل الإرسال.");
      var rb=document.getElementById("rcptw_"+id);
      if(rb){rb.classList.add("show");rb.scrollIntoView({behavior:"smooth",block:"center"});
        rb.style.boxShadow="0 0 0 3px rgba(192,57,43,.35)";
        setTimeout(function(){rb.style.boxShadow=""},1600);}
      return;
    }
  }
  if(st==="rejected"){
    var t=document.getElementById("rej_"+id);
    var reason=t?t.value.trim():"";
    if(!reason){if(t)t.focus();toast("سبب الرفض مطلوب قبل الإرسال.");return}
    payload.reject_reason=reason;
  }

  btn.disabled=true;btn.textContent="جارٍ الإرسال…";
  fetch(API+"/"+id+"/status",{method:"PATCH",credentials:"same-origin",headers:{"Content-Type":"application/json"},body:JSON.stringify(payload)})
    .then(function(r){if(!r.ok)return r.json().then(function(e){throw {msg:(e&&e.detail)||0}});return r.json()})
    .then(function(u){
      var msg;
      if(st==="done")msg="تم إتمام الإرجاع ✓ مع إيصال التحويل.";
      else if(st==="rejected")msg="تم رفض الطلب نهائياً وتسجيل السبب.";
      else msg="تم تحديث الحالة إلى: "+SL[st]+".";
      toast(msg);
      delete PICKED[id];
      var i=DATA.findIndex(function(d){return d.id===id});if(i>=0)DATA[i]=u;
      render();
    })
    .catch(function(e){
      toast((e&&e.msg)||"تعذّر تحديث الحالة، حاول مجدداً.");
      btn.disabled=false;btn.textContent="إرسال — "+SL[st];
    });
}
function initMe(cb){fetch("/returns/api/me",{credentials:"same-origin"}).then(function(r){if(r.status===401){location.href="/accountant-login";return null}return r.json()}).then(function(m){if(m){ME=m;applyRole();}if(cb)cb();}).catch(function(){if(cb)cb();});}
function applyRole(){
  // The "تم الإرجاع في سلة" tab is a follow-up view for EVERYONE:
  //  - النذير sees it to track what he confirmed;
  //  - the financial accountant sees it to pick up done_salla requests for the
  //    financial transfer step.
  // (Action button visibility is separate — driven per-status by canStatus(),
  //  so only النذير gets the done_salla *button*; the accountant only gets his
  //  four financial buttons.)
  var sallaTab=document.getElementById("tab-done_salla");
  if(sallaTab)sallaTab.style.display="";
  if(ME.is_purchaser){
    // النذير has no role in the financial pipeline: hide the financial status tabs.
    ["will","doing","done","rejected","old_done","old_rejected"].forEach(function(t){
      var b=document.querySelector('.tab[data-tab="'+t+'"]');
      if(b)b.style.display="none";
    });
    // default النذير to their own tab
    var own=document.getElementById("tab-done_salla");
    if(own)setTab("done_salla",own);
  }
}
initMe(function(){load();});
setInterval(load,20000);
</script></body></html>"""


# ------------------- Team "submitted requests" page -------------------
# Option B: NO bank account / IBAN columns (reduce spread of banking data).

# ── Team requests: shared filtering for list + export ──────────────────────────
# The list endpoint and the Excel export MUST read through this one function, so
# "what you see" and "what you export" can never drift apart.
#
# PRIVACY: iban / bank_account / refund_reference are deliberately NOT selected
# here. The team page only ever shows the refund METHOD; the full banking detail
# stays on the accountant page. The export inherits that boundary.
_TEAM_COLUMNS = """id, conversation_id, customer_name, order_number, reason,
                   status, accountant_note, reject_reason, assignee,
                   receipt_name, status_history, return_created_at, created_at,
                   updated_at, contacted_at, refund_method"""


def _parse_ymd(value, label):
    """'YYYY-MM-DD' -> datetime.date, or a 400 the caller can show the user."""
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{label} يجب أن يكون بصيغة YYYY-MM-DD")


async def _fetch_team_rows(date_from=None, date_to=None, status=None,
                           assignee=None, q=None, ids=None):
    """Filtered rows, newest first. No LIMIT — a silent cap would hide records
    from the exported sheet exactly the way it would hide them from the screen.

    `ids` wins over the other filters: the page sends the exact rows it is
    displaying, so "what you see" and "what you export" stay identical even for
    the قديمة tabs, whose age rule is computed in the browser from status_history.
    """
    where, args = [], []

    if ids:
        args.append(list(ids))
        where.append(f"id = ANY(${len(args)}::bigint[])")

    # asyncpg binds by Python type — a 'YYYY-MM-DD' string is rejected even with a
    # ::date cast in the SQL, so parse to a real date object before binding.
    if date_from:
        args.append(_parse_ymd(date_from, "date_from"))
        where.append(f"created_at::date >= ${len(args)}")
    if date_to:
        args.append(_parse_ymd(date_to, "date_to"))
        where.append(f"created_at::date <= ${len(args)}")
    if status:
        args.append(status)
        where.append(f"status = ${len(args)}")
    if assignee:
        args.append(assignee.strip())
        where.append(f"btrim(coalesce(assignee,'')) = ${len(args)}")
    if q and q.strip():
        args.append(f"%{q.strip()}%")
        where.append(
            f"(coalesce(customer_name,'') ILIKE ${len(args)} "
            f"OR coalesce(order_number,'') ILIKE ${len(args)} "
            f"OR coalesce(conversation_id::text,'') ILIKE ${len(args)})"
        )

    sql = f"SELECT {_TEAM_COLUMNS} FROM return_requests"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"

    p = await pool()
    async with p.acquire() as c:
        return await c.fetch(sql, *args)


@app.get("/returns/api/team-requests")
async def team_requests(
    response: Response,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    assignee: Optional[str] = None,
    q: Optional[str] = None,
):
    rows = await _fetch_team_rows(date_from, date_to, status, assignee, q)
    response.headers["X-Total-Count"] = str(len(rows))
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"
    out = []
    for r in rows:
        d = dict(r)
        for k in ("return_created_at",):
            if d.get(k) is not None:
                d[k] = d[k].isoformat()
        for k in ("created_at", "updated_at", "contacted_at"):
            if d.get(k) is not None:
                d[k] = d[k].isoformat()
        if isinstance(d.get("status_history"), str):
            try:
                d["status_history"] = json.loads(d["status_history"])
            except Exception:
                d["status_history"] = []
        d["status_label"] = STATUS_LABELS.get(d.get("status"), d.get("status"))
        # طريقة الاسترداد فقط — الرقم/المرجع والبيانات البنكية تبقى مستبعدة هنا (خصوصية).
        d["refund_method_label"] = REFUND_LABELS.get(d.get("refund_method") or "bank_account",
                                                    d.get("refund_method"))
        out.append(d)
    return out




# ── Excel export (team view) ───────────────────────────────────────────────────
# Same filters as the list endpoint, no row cap. Banking detail stays excluded:
# the sheet carries the refund METHOD only, matching what the team page shows.

_XL_FILL_TITLE = "1F5F5B"
_XL_FILL_META = "12403D"
_XL_FILL_HEADER = "2F4858"
_XL_FILL_TOTAL = "1F7A4D"
_XL_FILL_STRIPE = "F4F6F8"
_XL_FILL_REJECTED = "FDEDED"


def _fmt_dt(v, with_time=False):
    if not v:
        return "—"
    try:
        if with_time:
            return v.strftime("%Y-%m-%d %H:%M")
        return v.strftime("%Y-%m-%d")
    except Exception:
        return str(v)


def _build_returns_workbook(rows, period_label: str):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def fill(color):
        return PatternFill("solid", fgColor=color)

    border = Border(*[Side(style="thin", color="D8DEE4")] * 4)

    wb = Workbook()
    ws = wb.active
    ws.title = "طلبات الإرجاع"
    ws.sheet_view.rightToLeft = True

    cols = [
        ("#", 7), ("رقم المحادثة", 13), ("العميل", 22), ("رقم الطلب", 15),
        ("سبب الإرجاع", 26), ("طريقة الاسترداد", 16), ("الحالة", 16),
        ("ملاحظة المحاسب", 30), ("سبب الرفض", 26), ("الموظف", 12),
        ("إيصال التحويل", 12), ("تم التواصل", 14),
        ("تاريخ الإنشاء", 14), ("تاريخ إنشاء الإرجاع", 16), ("آخر تحديث", 16),
    ]
    ncols = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "تقرير طلبات الإرجاع — QAYDAO")
    c.font = Font(bold=True, size=15, color="FFFFFF")
    c.fill = fill(_XL_FILL_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, f"{period_label}   |   تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(size=10, color="D6E4E3")
    c.fill = fill(_XL_FILL_META)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    by_status, by_assignee, by_month = {}, {}, {}
    for r in rows:
        st = STATUS_LABELS.get(r["status"], r["status"] or "—")
        by_status[st] = by_status.get(st, 0) + 1
        ag = (r["assignee"] or "—").strip() or "—"
        by_assignee[ag] = by_assignee.get(ag, 0) + 1
        m = r["created_at"].strftime("%Y-%m") if r["created_at"] else "—"
        by_month[m] = by_month.get(m, 0) + 1

    kpis = [("إجمالي الطلبات", len(rows))] + [(k, v) for k, v in sorted(by_status.items(), key=lambda x: -x[1])][:4]
    span = max(1, ncols // max(1, len(kpis)))
    col = 1
    for label, value in kpis:
        last = min(col + span - 1, ncols)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=last)
        lc = ws.cell(3, col, label)
        lc.font = Font(size=9, color="A8CFCB")
        lc.fill = fill(_XL_FILL_META)
        lc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=last)
        vc = ws.cell(4, col, value)
        vc.font = Font(bold=True, size=12, color="FFFFFF")
        vc.fill = fill(_XL_FILL_META)
        vc.alignment = Alignment(horizontal="center")
        col = last + 1
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 22

    HROW = 5
    for ci, (title, width) in enumerate(cols, 1):
        cell = ws.cell(HROW, ci, title)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill(_XL_FILL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HROW].height = 28

    wrap_cols = {5, 8, 9}
    rr = HROW + 1
    for idx, r in enumerate(rows):
        is_rejected = r["status"] == "rejected"
        values = [
            r["id"],
            r["conversation_id"] if r["conversation_id"] is not None else "—",
            (r["customer_name"] or "—").strip(),
            (r["order_number"] or "—").strip(),
            (r["reason"] or "—").strip(),
            REFUND_LABELS.get(r["refund_method"] or "bank_account", r["refund_method"] or "—"),
            STATUS_LABELS.get(r["status"], r["status"] or "—"),
            (r["accountant_note"] or "—").strip(),
            (r["reject_reason"] or "—").strip(),
            (r["assignee"] or "—").strip() or "—",
            "نعم" if r["receipt_name"] else "لا",
            _fmt_dt(r["contacted_at"], True),
            _fmt_dt(r["created_at"]),
            _fmt_dt(r["return_created_at"]),
            _fmt_dt(r["updated_at"], True),
        ]
        row_fill = fill(_XL_FILL_REJECTED) if is_rejected else (fill(_XL_FILL_STRIPE) if idx % 2 else None)
        for ci, val in enumerate(values, 1):
            cell = ws.cell(rr, ci, val)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if ci in wrap_cols:
                cell.alignment = Alignment(horizontal="right", wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 7 and is_rejected:
                cell.font = Font(bold=True, color="C0392B")
        rr += 1

    tcell = ws.cell(rr, 1, "الإجمالي")
    tcell.font = Font(bold=True, color="FFFFFF")
    tcell.fill = fill(_XL_FILL_TOTAL)
    tcell.alignment = Alignment(horizontal="center")
    for ci in range(2, ncols + 1):
        cell = ws.cell(rr, ci, f"{len(rows)} طلب" if ci == 2 else "")
        cell.fill = fill(_XL_FILL_TOTAL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    ws.row_dimensions[rr].height = 24

    note = ws.cell(rr + 2, 1,
                   "ملاحظة: البيانات البنكية (الآيبان ورقم الحساب ومرجع الاسترداد) مستبعدة عمداً من هذا التقرير — "
                   "تظهر في صفحة المحاسب فقط. الصفوف المرفوضة بخلفية حمراء.")
    note.font = Font(size=9, color="5A6B7D")
    ws.merge_cells(start_row=rr + 2, start_column=1, end_row=rr + 2, end_column=ncols)

    ws.freeze_panes = ws.cell(HROW + 1, 1)
    ws.auto_filter.ref = f"A{HROW}:{get_column_letter(ncols)}{rr - 1}"

    # ── Sheet 2: breakdown ──
    ws2 = wb.create_sheet("الملخص")
    ws2.sheet_view.rightToLeft = True
    for letter, width in (("A", 26), ("B", 14)):
        ws2.column_dimensions[letter].width = width

    def block(start_row, title, data):
        ws2.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        h = ws2.cell(start_row, 1, title)
        h.font = Font(bold=True, size=12, color="FFFFFF")
        h.fill = fill(_XL_FILL_TITLE)
        h.alignment = Alignment(horizontal="center")
        r = start_row + 1
        for ci, t in enumerate(["البند", "عدد الطلبات"], 1):
            hc = ws2.cell(r, ci, t)
            hc.font = Font(bold=True, color="FFFFFF")
            hc.fill = fill(_XL_FILL_HEADER)
            hc.alignment = Alignment(horizontal="center")
            hc.border = border
        r += 1
        for k, v in sorted(data.items(), key=lambda x: -x[1]):
            ws2.cell(r, 1, k).border = border
            vc = ws2.cell(r, 2, v)
            vc.alignment = Alignment(horizontal="center")
            vc.border = border
            r += 1
        tc = ws2.cell(r, 1, "الإجمالي")
        tc.font = Font(bold=True, color="FFFFFF")
        tc.fill = fill(_XL_FILL_TOTAL)
        tc.border = border
        vc = ws2.cell(r, 2, sum(data.values()))
        vc.font = Font(bold=True, color="FFFFFF")
        vc.fill = fill(_XL_FILL_TOTAL)
        vc.alignment = Alignment(horizontal="center")
        vc.border = border
        return r + 3

    nxt = block(1, "حسب الحالة", by_status)
    nxt = block(nxt, "حسب الموظف", by_assignee)
    block(nxt, "حسب الشهر", by_month)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/returns/api/team-requests/export")
async def team_requests_export(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    assignee: Optional[str] = None,
    q: Optional[str] = None,
    ids: Optional[str] = None,
    scope: Optional[str] = None,
):
    """Excel of every row matching the filters — deliberately uncapped."""
    for label, val in (("date_from", date_from), ("date_to", date_to)):
        if val and not re.match(r"^\d{4}-\d{2}-\d{2}$", val):
            raise HTTPException(status_code=400, detail=f"{label} يجب أن يكون بصيغة YYYY-MM-DD")
    if status and status not in STATUS_LABELS:
        raise HTTPException(status_code=400, detail="حالة غير معروفة")

    id_list = None
    if ids and ids.strip():
        try:
            id_list = [int(x) for x in ids.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="قائمة المعرّفات غير صالحة")
        if not id_list:
            id_list = None

    rows = await _fetch_team_rows(date_from, date_to, status, assignee, q, id_list)

    parts = []
    if date_from or date_to:
        parts.append(f"الفترة: {date_from or 'البداية'} → {date_to or 'اليوم'}")
    else:
        parts.append("الفترة: كل الطلبات")
    if status:
        parts.append(f"الحالة: {STATUS_LABELS.get(status, status)}")
    if assignee:
        parts.append(f"الموظف: {assignee}")
    if q and q.strip():
        parts.append(f"بحث: {q.strip()}")
    if scope and scope.strip():
        parts.append(f"العرض: {scope.strip()}")
    period_label = "  |  ".join(parts)

    buf = _build_returns_workbook(rows, period_label)
    fname = f"returns_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={fname}",
            "X-Total-Count": str(len(rows)),
        },
    )

@app.get("/returns/team-requests", response_class=HTMLResponse)
async def team_page():
    return HTMLResponse(TEAM_HTML)


TEAM_HTML = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>الطلبات المرفوعة — المرجعات — QAYDAO</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{--bg:#f4f6f8;--surface:#fff;--ink:#1f2b3a;--soft:#5a6b7d;--line:#e4e9ee;--brand:#1f5f5b;--brandsoft:#e8f1f0;--ok:#1f7a4d;--oksoft:#e6f4ec;--amber:#b5791d;--ambersoft:#fbf0d9;--info:#2c5a86;--infosoft:#eef4fb;--rej:#c0392b;--rejsoft:#fdeded}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Cairo,system-ui,sans-serif;background:var(--bg);color:var(--ink);line-height:1.6;padding-bottom:50px}
.ltr{direction:ltr;unicode-bidi:isolate;display:inline-block}
.wrap{max-width:1240px;margin:0 auto;padding:0 18px}
.topbar{background:#12403d;color:#fff;text-align:center;padding:8px;font-size:12.5px;font-weight:600}
header{background:var(--surface);border-bottom:1px solid var(--line);padding:18px 0}
.hrow{display:flex;align-items:center;gap:13px;flex-wrap:wrap}
.logo{width:42px;height:42px;border-radius:11px;background:var(--brand);color:#fff;display:grid;place-items:center;font-weight:800;font-size:18px}
h1{font-size:19px;font-weight:800}
header p{font-size:12.5px;color:var(--soft);margin-top:2px}
.motiv{margin-inline-start:auto;max-width:290px;text-align:start;background:linear-gradient(135deg,var(--brandsoft),#f3f9f8);border:1px solid #d5e6e4;border-radius:12px;padding:11px 15px;display:none}
.motiv.show{display:block}
.motiv .mn{font-size:13.5px;font-weight:800;color:var(--brand)}
.motiv .mp{font-size:12px;color:#2b4a47;margin-top:2px;line-height:1.5}
.tools{display:flex;gap:10px;align-items:center;margin:16px 0;flex-wrap:wrap}
.agents{display:flex;gap:11px;flex-wrap:wrap;margin:16px 0 4px}
.tabs{display:flex;gap:8px;flex-wrap:wrap;margin:14px 0 4px}
.tab{font-family:inherit;font-size:12.5px;font-weight:700;cursor:pointer;border:1px solid var(--line);background:#fff;color:var(--soft);border-radius:999px;padding:7px 14px;transition:.15s;display:flex;align-items:center;gap:6px}
.tab:hover{border-color:var(--brand);color:var(--brand)}
.tab.active{background:var(--brand);color:#fff;border-color:var(--brand)}
.tab .cnt{font-size:10.5px;font-weight:700;background:rgba(0,0,0,.12);border-radius:999px;padding:1px 6px;min-width:16px;text-align:center}
.tab.active .cnt{background:rgba(255,255,255,.28)}
.tab-old{border-style:dashed;color:#8a92a0}
.tab-old:hover{border-color:#8a92a0;color:#5a6b7d}
.tab-old.active{background:#5a6b7d;border-color:#5a6b7d;border-style:solid;color:#fff}
.agent{cursor:pointer;background:#fff;border:1.5px solid var(--line);border-radius:14px;padding:13px 20px;min-width:112px;text-align:center;transition:.15s;font-family:inherit}
.agent:hover{border-color:var(--brand);transform:translateY(-1px);box-shadow:0 4px 14px rgba(31,43,58,.08)}
.agent.active{background:var(--brand);border-color:var(--brand);color:#fff;box-shadow:0 4px 14px rgba(31,95,91,.25)}
.agent .an{font-size:14px;font-weight:800;display:block}
.agent .ac{font-size:11.5px;font-weight:600;color:var(--soft);margin-top:3px;display:block}
.agent.active .ac{color:rgba(255,255,255,.85)}
.agent .arej{display:inline-block;font-size:10px;font-weight:700;color:var(--rej);background:var(--rejsoft);border-radius:999px;padding:1px 7px;margin-top:4px}
.agent.active .arej{background:rgba(255,255,255,.22);color:#fff}
.tools select,.tools input{font-family:inherit;font-size:13.5px;padding:8px 12px;border:1px solid var(--line);border-radius:10px;background:#fff}
.refresh{background:var(--brand);color:#fff;border:none;border-radius:10px;padding:8px 15px;font-family:inherit;font-weight:700;font-size:13px;cursor:pointer}
.dategrp{display:flex;gap:6px;align-items:center;font-size:12.5px;color:var(--soft)}
.dategrp input[type=date]{padding:7px 10px}
.clearf{background:#fff;color:var(--soft);border:1px solid var(--line);border-radius:10px;padding:8px 12px;font-family:inherit;font-size:12.5px;cursor:pointer}
.clearf:hover{background:#f4f6f8}
.xbtn{margin-inline-start:auto;background:#1f7a4d;color:#fff;border:none;border-radius:10px;padding:8px 15px;font-family:inherit;font-weight:700;font-size:13px;cursor:pointer;white-space:nowrap}
.xbtn:hover{background:#19643f}
.xbtn:disabled{opacity:.55;cursor:not-allowed}
.count{font-size:12.5px;color:var(--soft)}
.rejbar{background:var(--rejsoft);border:1px solid #f3c6c1;color:var(--rej);border-radius:12px;padding:12px 15px;margin-bottom:14px;font-size:13px;font-weight:600;display:none}
.rejbar.show{display:block}
.rejbar a{color:var(--rej);font-weight:800}
.tablewrap{background:var(--surface);border:1px solid var(--line);border-radius:14px;overflow:hidden;box-shadow:0 1px 2px rgba(31,43,58,.04),0 6px 20px rgba(31,43,58,.05)}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th,td{padding:11px 12px;text-align:start;border-bottom:1px solid #eef1f4;vertical-align:top}
th{background:#f8fafb;font-weight:700;color:var(--soft);white-space:nowrap;font-size:12px}
tr:last-child td{border-bottom:none}
tr.rejrow{background:var(--rejsoft)}
.badge{font-size:11px;font-weight:700;padding:3px 9px;border-radius:999px;white-space:nowrap}
.b-new{background:#eef1f4;color:#5a6b7d}.b-will{background:var(--ambersoft);color:var(--amber)}
.b-doing{background:var(--infosoft);color:var(--info)}.b-done{background:var(--oksoft);color:var(--ok)}
.b-rejected{background:var(--rejsoft);color:var(--rej)}
.b-done_salla{background:#eef4ff;color:#3b5bdb}
.b-rejected_contacted{background:#eef1f4;color:#5a6b7d;border:1px solid #d7dde4}
tr.contactedrow{background:#fbfcfd}
.contact-btn{font-family:inherit;font-size:11px;font-weight:700;cursor:pointer;border-radius:8px;padding:4px 10px;background:#fff;color:var(--rej);border:1px solid #f3c6c1;transition:.15s;white-space:nowrap}
.contact-btn:hover{background:var(--rejsoft)}
.contact-done{font-size:10.5px;font-weight:700;color:var(--soft);white-space:nowrap}
.note{color:var(--ink)}.rej{color:var(--rej);font-weight:700}
.rcpt-dl{color:var(--ok);font-weight:700;text-decoration:none;white-space:nowrap;background:var(--oksoft);border:1px solid #bfe3cd;border-radius:7px;padding:3px 9px;font-size:11.5px;display:inline-block}
.rcpt-dl:hover{background:#d9f0e2}
.empty{text-align:center;color:var(--soft);padding:50px 20px;font-size:15px;display:none}
footer{text-align:center;margin-top:22px;font-size:11.5px;color:var(--soft)}
@media(max-width:720px){th:nth-child(3),td:nth-child(3){display:none}}
</style></head><body>
<div class="topbar">📋 الطلبات المرفوعة — متابعة ردّ المحاسب على طلبات الإرجاع</div>
<header><div class="wrap hrow">
  <div class="logo">Q</div>
  <div><h1>الطلبات المرفوعة للمحاسب</h1><p>تابع حالة كل طلب إرجاع رفعته للمحاسب وردّه عليه — للرد على العميل.</p></div>
  <div class="motiv" id="motiv"></div>
</div></header>
<div class="wrap">
  <div id="rejbar" class="rejbar"></div>
  <div class="agents" id="agents"></div>
  <div class="tabs" id="tabs">
    <button class="tab active" data-tab="all" onclick="setTab('all',this)">الكل <span class="cnt" id="tcnt-all">0</span></button>
    <button class="tab" data-tab="new" onclick="setTab('new',this)">جديدة <span class="cnt" id="tcnt-new">0</span></button>
    <button class="tab" data-tab="will" onclick="setTab('will',this)">سيتم الإرجاع <span class="cnt" id="tcnt-will">0</span></button>
    <button class="tab" data-tab="doing" onclick="setTab('doing',this)">جاري الإرجاع <span class="cnt" id="tcnt-doing">0</span></button>
    <button class="tab" data-tab="done" onclick="setTab('done',this)">تم الإرجاع <span class="cnt" id="tcnt-done">0</span></button>
    <button class="tab" data-tab="rejected" onclick="setTab('rejected',this)">مرفوض <span class="cnt" id="tcnt-rejected">0</span></button>
    <button class="tab tab-old" data-tab="old_done" onclick="setTab('old_done',this)">قديمة — تم الإرجاع <span class="cnt" id="tcnt-old_done">0</span></button>
    <button class="tab tab-old" data-tab="old_rejected" onclick="setTab('old_rejected',this)">قديمة — مرفوضة <span class="cnt" id="tcnt-old_rejected">0</span></button>
  </div>
  <div class="tools">
    <span class="dategrp">
      من <input type="date" id="fdate_from" onchange="render()">
      إلى <input type="date" id="fdate_to" onchange="render()">
    </span>
    <input id="fsearch" placeholder="بحث بالاسم أو رقم المحادثة أو الطلب…" oninput="render()">
    <button class="clearf" onclick="clearFilters()">مسح الفلاتر</button>
    <span class="count" id="count"></span>
    <button class="refresh" onclick="load()">تحديث ⟳</button>
    <button class="xbtn" id="xbtn" onclick="exportExcel()" title="تصدير الطلبات الظاهرة حالياً إلى ملف Excel">📥 تصدير Excel</button>
  </div>
  <div class="tablewrap">
    <table>
      <thead><tr>
        <th>رقم المحادثة</th><th>العميل</th><th>رقم الطلب</th><th>سبب الإرجاع</th>
        <th>طريقة الاسترداد</th><th>الحالة</th><th>ملاحظة المحاسب</th><th>سبب الرفض</th><th>إيصال التحويل</th><th>الموظف</th><th>إجراء</th>
      </tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
  <div class="empty" id="empty">لا توجد طلبات مرفوعة بعد.</div>
  <footer>QAYDAO · الطلبات المرفوعة · البيانات البنكية غير معروضة هنا لدواعي الخصوصية — تظهر في نموذج الإدخال فقط</footer>
</div>
<script>
var API="/returns/api/team-requests";
var EMAIL="financial@qaydao.com";
(function(){
  function param(n){try{return new URLSearchParams(location.search).get(n)||""}catch(e){return ""}}
  function firstName(f){f=(f||"").trim();return f?f.split(/\s+/)[0]:""}
  function esc0(s){return (s==null?"":String(s)).replace(/[&<>"]/g,function(c){return{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]})}
  var nm=firstName(param("agent"));
  var phrase="ابتسامتك هي قوتك وأداؤك هو نجاحك";
  var box=document.getElementById("motiv");
  if(box){
    box.innerHTML=(nm?('<div class="mn">\u2728 '+esc0(nm)+'</div>'):'<div class="mn">\u2728 أهلاً بك</div>')+
      '<div class="mp">'+esc0(phrase)+'</div>';
    box.classList.add("show");
  }
})();
var DATA=[];
var SL={new:"جديد",will:"سيتم الإرجاع",doing:"جاري الإرجاع",done:"تم الإرجاع",rejected:"مرفوض",done_salla:"تم الإرجاع في سلة"};
function esc(s){return (s==null?"":String(s)).replace(/[&<>"]/g,function(c){return{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]})}
function load(){fetch(API).then(function(r){return r.json()}).then(function(d){DATA=Array.isArray(d)?d:[];render()}).catch(function(){})}
var CURAGENT="";  // "" = الكل
function setAgent(a){CURAGENT=a;render()}

/* --- status sections (same rules as the accountant page) --- */
var CURTAB="all";
var OLD_DAYS=7;
function closedAt(x){
  var h=x.status_history;
  if(h&&h.length){var last=h[h.length-1];if(last&&last.at)return last.at}
  return x.updated_at||x.return_created_at;
}
function isOld(x){
  var t=closedAt(x);if(!t)return false;
  var d=new Date(t);if(isNaN(d))return false;
  return (Date.now()-d.getTime())>OLD_DAYS*86400000;
}
function inTab(x){
  switch(CURTAB){
    case "all":          return true;
    case "new":          return x.status==="new";
    case "will":         return x.status==="will";
    case "doing":        return x.status==="doing";
    case "done":         return x.status==="done"     && !isOld(x);
    case "rejected":     return x.status==="rejected" && !isOld(x);
    case "old_done":     return x.status==="done"     &&  isOld(x);
    case "old_rejected": return x.status==="rejected" &&  isOld(x);
    default:             return true;
  }
}
function setTab(t,btn){
  CURTAB=t;
  var tabs=document.querySelectorAll("#tabs .tab");
  for(var i=0;i<tabs.length;i++)tabs[i].classList.remove("active");
  if(btn)btn.classList.add("active");
  render();
}
function updateTabCounts(scope){
  var c={all:scope.length,new:0,will:0,doing:0,done:0,rejected:0,old_done:0,old_rejected:0};
  scope.forEach(function(x){
    var o=isOld(x);
    if(x.status==="new")c.new++;
    else if(x.status==="will")c.will++;
    else if(x.status==="doing")c.doing++;
    else if(x.status==="done"){o?c.old_done++:c.done++}
    else if(x.status==="rejected"){o?c.old_rejected++:c.rejected++}
  });
  Object.keys(c).forEach(function(k){
    var e=document.getElementById("tcnt-"+k);if(e)e.textContent=c[k];
  });
}

/* agent list: from Chatwoot (?agents=) + anyone appearing in the data */
function agentNames(){
  var names=[];
  try{
    var p=new URLSearchParams(location.search).get("agents");
    if(p)p.split("|").forEach(function(n){n=n.trim();if(n&&names.indexOf(n)<0)names.push(n)});
  }catch(e){}
  DATA.forEach(function(x){
    var a=(x.assignee||"").trim();
    if(a&&names.indexOf(a)<0)names.push(a);
  });
  return names;
}
function renderAgents(){
  var names=agentNames();
  var box=document.getElementById("agents");
  if(!box)return;
  var html='<button class="agent'+(CURAGENT===""?" active":"")+'" onclick="setAgent(\'\')">'+
    '<span class="an">الكل</span><span class="ac">'+DATA.length+' طلب</span></button>';
  names.forEach(function(n){
    var mine=DATA.filter(function(x){return (x.assignee||"").trim()===n});
    var rej=mine.filter(function(x){return x.status==="rejected"}).length;
    html+='<button class="agent'+(CURAGENT===n?" active":"")+'" onclick="setAgent(\''+esc(n).replace(/'/g,"\\'")+'\')">'+
      '<span class="an">'+esc(n)+'</span>'+
      '<span class="ac">'+mine.length+' طلب</span>'+
      (rej?'<span class="arej">'+rej+' مرفوض</span>':'')+
    '</button>';
  });
  box.innerHTML=html;
}
var VISIBLE=[];
// Date filter runs on created_at (when the request was raised), which is the
// field the accountant reconciles against.
function inDateRange(x,df,dt){
  var raw=x.created_at||x.return_created_at;
  if(!raw)return !(df||dt);
  var d=String(raw).slice(0,10);
  if(df && d<df)return false;
  if(dt && d>dt)return false;
  return true;
}
function clearFilters(){
  document.getElementById("fdate_from").value="";
  document.getElementById("fdate_to").value="";
  document.getElementById("fsearch").value="";
  render();
}
function tabLabel(){
  var b=document.querySelector('#tabs .tab[data-tab="'+CURTAB+'"]');
  if(!b)return "";
  return b.textContent.replace(/\s*\d+\s*$/,"").trim();
}
function exportExcel(){
  var btn=document.getElementById("xbtn");
  if(!VISIBLE.length){alert("لا توجد طلبات مطابقة للفلاتر الحالية.");return}
  btn.disabled=true;var old=btn.textContent;btn.textContent="جارٍ التصدير…";
  var p=new URLSearchParams();
  p.set("ids",VISIBLE.map(function(x){return x.id}).join(","));
  var df=document.getElementById("fdate_from").value;
  var dt=document.getElementById("fdate_to").value;
  if(df)p.set("date_from",df);
  if(dt)p.set("date_to",dt);
  var qq=document.getElementById("fsearch").value.trim();
  if(qq)p.set("q",qq);
  if(CURAGENT)p.set("assignee",CURAGENT);
  var sc=tabLabel();if(sc)p.set("scope",sc);
  fetch(API+"/export?"+p.toString())
    .then(function(r){if(!r.ok)throw new Error("HTTP "+r.status);return r.blob()})
    .then(function(b){
      var u=URL.createObjectURL(b),a=document.createElement("a");
      a.href=u;a.download="returns_"+new Date().toISOString().slice(0,10)+".xlsx";
      document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(u);
    })
    .catch(function(e){alert("تعذّر التصدير: "+e.message)})
    .finally(function(){btn.disabled=false;btn.textContent=old});
}
function render(){
  renderAgents();
  // agent scope first, then the status tab, then search
  var scope=DATA.filter(function(x){
    return !CURAGENT || (x.assignee||"").trim()===CURAGENT;
  });
  updateTabCounts(scope);
  var q=document.getElementById("fsearch").value.trim().toLowerCase();
  var df=document.getElementById("fdate_from").value;
  var dt=document.getElementById("fdate_to").value;
  var list=scope.filter(function(x){
    if(!inTab(x))return false;
    if(!inDateRange(x,df,dt))return false;
    if(q){var h=((x.customer_name||"")+" "+(x.conversation_id||"")+" "+(x.order_number||"")).toLowerCase();if(h.indexOf(q)<0)return false}
    return true;
  });
  // Keep the exact rows on screen so the Excel export can send their ids and
  // reproduce this view byte-for-byte instead of re-deriving the filters.
  VISIBLE=list;
  var totalTxt=list.length+" طلب";
  if(list.length!==DATA.length){totalTxt+=" من أصل "+DATA.length}
  document.getElementById("count").textContent=totalTxt;
  // Alert counts ONLY rejected requests the agent has NOT yet followed up on.
  // Pressing "تم التواصل" sets contacted_at → the request drops out of this count.
  var rejected=list.filter(function(x){return x.status==="rejected" && !x.contacted_at});
  var rb=document.getElementById("rejbar");
  if(rejected.length){rb.classList.add("show");rb.innerHTML="\u26A0 يوجد "+rejected.length+" طلب مرفوض. لمعرفة تفاصيل سبب الرفض تواصل مع المحاسب عبر الإيميل: <a href=\"mailto:"+EMAIL+"\">"+EMAIL+"</a>"}
  else{rb.classList.remove("show");rb.innerHTML=""}
  var tb=document.getElementById("tbody"),e=document.getElementById("empty");
  if(!list.length){tb.innerHTML="";e.style.display="block";return}
  e.style.display="none";
  tb.innerHTML=list.map(row).join("");
}
function row(x){
  var rej=x.status==="rejected";
  var contacted=rej && !!x.contacted_at;          // rejected AND already followed up
  var openRej=rej && !x.contacted_at;             // rejected, still needs follow-up (red)
  // status badge: a contacted-rejected request gets the calm "مرفوض — تم التواصل" badge
  var badge = contacted
    ? '<span class="badge b-rejected_contacted">مرفوض — تم التواصل</span>'
    : '<span class="badge b-'+esc(x.status)+'">'+esc(SL[x.status]||x.status)+'</span>';
  // reject-reason cell: keep the "email the accountant" hint only while still open
  var rejCell = rej
    ? (esc(x.reject_reason||"—") + (openRej ? '<br><span style="font-size:11px">\u26A0 تواصل مع المحاسب بالإيميل</span>' : ''))
    : '—';
  // action cell: "تم التواصل" button only for still-open rejected requests
  var action='—';
  if(openRej){
    action='<button class="contact-btn" onclick="markContacted('+x.id+',this)">\u2713 تم التواصل</button>';
  }else if(contacted){
    action='<span class="contact-done">\u2713 تمت المتابعة</span>';
  }
  return '<tr class="'+(openRej?"rejrow":(contacted?"contactedrow":""))+'">'+
    '<td class="ltr">'+esc(x.conversation_id||"—")+'</td>'+
    '<td>'+esc(x.customer_name||"—")+'</td>'+
    '<td class="ltr">'+esc(x.order_number||"—")+'</td>'+
    '<td>'+esc(x.reason||"—")+'</td>'+
    '<td>'+esc(x.refund_method_label||"—")+'</td>'+
    '<td>'+badge+'</td>'+
    '<td class="note">'+esc(x.accountant_note||"—")+'</td>'+
    '<td class="'+(openRej?"rej":"")+'">'+rejCell+'</td>'+
    '<td>'+(x.receipt_name?('<a class="rcpt-dl" href="/returns/api/requests/'+x.id+'/receipt" target="_blank">\u2B07 تحميل الإيصال</a>'):'—')+'</td>'+
    '<td>'+esc(x.assignee||"—")+'</td>'+
    '<td>'+action+'</td>'+
  '</tr>';
}
function markContacted(id,btn){
  btn.disabled=true;btn.textContent="جارٍ…";
  fetch(API.replace("/team-requests","")+"/requests/"+id+"/contacted",{method:"POST"})
    .then(function(r){if(!r.ok)return r.json().then(function(e){throw {msg:(e&&e.detail)||"تعذّر التسجيل"}});return r.json()})
    .then(function(u){
      var i=DATA.findIndex(function(d){return d.id===id});if(i>=0)DATA[i].contacted_at=u.contacted_at;
      render();
    })
    .catch(function(e){btn.disabled=false;btn.textContent="\u2713 تم التواصل";alert((e&&e.msg)||"تعذّر تسجيل التواصل، حاول مجدداً.")});
}
load();
setInterval(load,20000);
</script></body></html>"""
